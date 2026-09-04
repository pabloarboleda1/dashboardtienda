"""
====================================================================
 REPORTE DE VENTAS - TIENDA VIRTUAL UCN
====================================================================
Dashboard en Streamlit para analizar y comparar los archivos
mensuales "Control_Integral_Inventario_<MES>.xlsx" de la Tienda
Virtual UCN.

Usa dos hojas de cada archivo:
    - "Consignaciones": datos por pedido (canal, valor, descripción).
    - "Inventario":      catálogo de productos, origen de compra y
                          ventas por canal (Física/Online/Nómina).
    - "Retribución FONDO": tabla de tasas de retribución por categoría.

Funciones principales:
    - KPIs, evolución y comparación de ingresos por periodo/canal.
    - Top 5 productos más vendidos.
    - Control de calidad: productos de Consignaciones sin coincidencia
      en Inventario.
    - Cálculo de la Retribución al Fondo de Empleados.
    - Descarga de un Excel de "Análisis de Fin de Mes".

Cómo ejecutar:
    streamlit run app.py
====================================================================
"""

import io
import re
import unicodedata
import zipfile
from datetime import datetime

import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from groq import Groq
    GROQ_DISPONIBLE = True
except ImportError:
    GROQ_DISPONIBLE = False

MODELO_GROQ = "openai/gpt-oss-120b"  # llama-3.3-70b-versatile fue descontinuado por Groq en jun/2026

# --------------------------------------------------------------
# CONFIGURACIÓN GENERAL
# --------------------------------------------------------------
st.set_page_config(
    page_title="Reporte de Ventas | Tienda Virtual UCN",
    page_icon="chart",
    layout="wide",
    initial_sidebar_state="expanded",
)

COLUMNAS_EXCEL_CONSIGNACIONES = "F,G,I,J,N,O,P"
FILA_ENCABEZADO_CONSIGNACIONES = 1  # 0-indexado -> fila 2 de Excel
COLUMNAS_INTERNAS_CONSIGNACIONES = [
    "Retribucion", "Medio_Pago", "Valor_Mercancia", "Envio",
    "Consignacion_Neto", "Fecha_Pago", "Descripcion",
]

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}
MESES_ALIAS_ARCHIVO = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "SETIEMBRE": 9, "OCTUBRE": 10,
    "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

# Correcciones para encabezados que a veces se guardan con tildes corruptas
ALIAS_ENCABEZADOS_INVENTARIO = {
    "cantidadvendidafsica": "cantidadvendidafisica",
    "cantdeduccinnmina": "cantdeduccionnomina",
}
ALIAS_ENCABEZADOS_CONSIGNACIONES = {
    "retribucin": "retribucion",
    "consignacin": "consignacion",
    "comisin": "comision",
}

# --------------------------------------------------------------
# TEMAS (claro / oscuro)
# --------------------------------------------------------------
TEMAS = {
    "Claro": {
        "bg_app": "#EEF2F8", "bg_sidebar": "#FFFFFF", "bg_card": "#FFFFFF",
        "texto": "#0F172A", "texto_muted": "#5B6B82", "borde": "#E1E7F0",
        "plotly_template": "plotly_white", "chart_bg": "#FFFFFF",
    },
    "Oscuro": {
        "bg_app": "#0B1220", "bg_sidebar": "#111A2C", "bg_card": "#161F33",
        "texto": "#F1F5F9", "texto_muted": "#9AA8C0", "borde": "#26324A",
        "plotly_template": "plotly_dark", "chart_bg": "#161F33",
    },
}
ACENTO_INGRESOS = "#2F6FED"
ACENTO_BRUTO = "#8B5CF6"
ACENTO_UNIDADES = "#15B36F"
ACENTO_PEDIDOS = "#F5A524"
ACENTO_ALERTA = "#E5484D"
ACENTO_RETRIBUCION = "#F5A524"
PALETA_CATEGORIAS = ["#2F6FED", "#F5A524", "#15B36F", "#E5484D", "#8B5CF6", "#0EA5B7", "#F45D9C"]

# Colores para el Excel exportado (mismo espíritu que la planilla original del usuario)
XL_COLOR_HEADER = "2F4F8F"
XL_COLOR_SUBTOTAL = "D9E1F2"
XL_COLOR_OK = "C6EFCE"
XL_COLOR_ALERT = "FFEB9C"
XL_FUENTE = "Arial"


def inyectar_estilos(tema):
    st.markdown(
        f"""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
        html, body, [class*="css"] {{ font-family: 'Inter', -apple-system, sans-serif; }}

        #MainMenu {{visibility: hidden;}}
        footer {{visibility: hidden;}}
        .stDeployButton {{display: none;}}
        header[data-testid="stHeader"] {{background: transparent;}}

        [data-testid="stAppViewContainer"] {{ background-color: {tema['bg_app']}; }}
        [data-testid="stSidebar"] {{ background-color: {tema['bg_sidebar']}; border-right: 1px solid {tema['borde']}; }}
        .block-container {{ padding-top: 1.4rem; max-width: 1250px; }}
        h1, h2, h3, p, span, label, .stMarkdown {{ color: {tema['texto']}; }}

        .app-header {{ padding-bottom: 1rem; margin-bottom: 1.4rem; border-bottom: 2px solid {tema['borde']}; }}
        .app-header .eyebrow {{ font-size: 0.78rem; font-weight: 700; letter-spacing: 0.12em; text-transform: uppercase; color: {ACENTO_INGRESOS}; }}
        .app-header .title {{ font-size: 1.85rem; font-weight: 800; color: {tema['texto']}; line-height: 1.2; }}
        .app-header .subtitle {{ font-size: 0.9rem; color: {tema['texto_muted']}; margin-top: 0.15rem; }}

        .sidebar-brand {{ padding: 0.1rem 0 1rem 0; margin-bottom: 0.8rem; border-bottom: 1px solid {tema['borde']}; }}
        .sidebar-brand .name {{ font-size: 1.02rem; font-weight: 800; color: {tema['texto']}; }}
        .sidebar-brand .tag {{ font-size: 0.76rem; color: {tema['texto_muted']}; }}
        .sidebar-section {{ font-size: 0.72rem; font-weight: 700; letter-spacing: 0.08em; text-transform: uppercase; color: {tema['texto_muted']}; margin: 1rem 0 0.4rem 0; }}

        .kpi-card {{ background: {tema['bg_card']}; border-radius: 12px; border: 1px solid {tema['borde']}; padding: 1.05rem 1.25rem; box-shadow: 0 2px 6px rgba(0,0,0,0.08); }}
        .kpi-card .kpi-label {{ font-size: 0.76rem; font-weight: 700; letter-spacing: 0.04em; text-transform: uppercase; color: {tema['texto_muted']}; margin-bottom: 0.4rem; }}
        .kpi-card .kpi-value {{ font-size: 1.55rem; font-weight: 800; }}

        .section-title {{ font-size: 1.05rem; font-weight: 700; color: {tema['texto']}; margin: 0.2rem 0 0.8rem 0; }}
        .section-caption {{ font-size: 0.82rem; color: {tema['texto_muted']}; margin: -0.5rem 0 0.8rem 0; }}
        [data-testid="stDataFrame"] {{ border: 1px solid {tema['borde']}; border-radius: 8px; }}

        .qa-ok {{ background: rgba(21, 179, 111, 0.12); border: 1px solid {ACENTO_UNIDADES}; color: {tema['texto']}; padding: 0.65rem 1rem; border-radius: 8px; font-size: 0.86rem; margin-bottom: 0.6rem; }}
        .qa-alert {{ background: rgba(229, 72, 77, 0.12); border: 1px solid {ACENTO_ALERTA}; color: {tema['texto']}; padding: 0.65rem 1rem; border-radius: 8px; font-size: 0.86rem; margin-bottom: 0.6rem; }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_kpi(col, etiqueta, valor, color_valor):
    col.markdown(
        f"""<div class="kpi-card"><div class="kpi-label">{etiqueta}</div>
        <div class="kpi-value" style="color:{color_valor};">{valor}</div></div>""",
        unsafe_allow_html=True,
    )


# --------------------------------------------------------------
# UTILIDADES DE TEXTO / NÚMEROS
# --------------------------------------------------------------
def normalizar_texto(texto):
    texto = str(texto).strip().lower()
    texto = "".join(c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn")
    texto = re.sub(r"[^a-z0-9]", "", texto)
    return texto


def limpiar_valor_numerico(valor):
    if pd.isna(valor):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = re.sub(r"[^\d,.\-]", "", str(valor))
    if texto == "":
        return None
    tiene_coma, tiene_punto = "," in texto, "." in texto
    if tiene_coma and tiene_punto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif tiene_coma:
        partes = texto.split(",")
        texto = texto.replace(",", "") if len(partes) > 1 and len(partes[-1]) == 3 else texto.replace(",", ".")
    elif tiene_punto:
        partes = texto.split(".")
        if len(partes) > 1 and len(partes[-1]) == 3:
            texto = texto.replace(".", "")
    try:
        return float(texto)
    except ValueError:
        return None


def formato_pesos(valor):
    signo = "-" if valor < 0 else ""
    return f"{signo}$ {abs(valor):,.0f}".replace(",", ".")


def construir_lookup(columnas, alias_dict):
    lookup = {}
    for col in columnas:
        norm = normalizar_texto(col)
        norm = alias_dict.get(norm, norm)
        lookup[norm] = col
    return lookup


def obtener_columna(lookup, *candidatos):
    for c in candidatos:
        if c in lookup:
            return lookup[c]
    return None


def categoria_retribucion(nombre_producto):
    n = str(nombre_producto).upper()
    if "CAMISETA" in n:
        return "CAMISETAS"
    if "CHAQUETAC" in n or "CHAQUETA" in n:
        return "CHAQUETAS CORTAVIENTOS"
    if "CHOMPA" in n:
        return "CHAQUETAS ABULLONADAS"
    if "GORRA" in n:
        return "GORRAS"
    return None


def detectar_hoja(xls, palabra_clave):
    for nombre in xls.sheet_names:
        if palabra_clave in nombre.strip().lower():
            return nombre
    return None


def detectar_periodo(nombre_archivo, serie_fechas):
    nombre_upper = nombre_archivo.upper()
    mes_num = None
    for nombre_mes, num in MESES_ALIAS_ARCHIVO.items():
        if nombre_mes in nombre_upper:
            mes_num = num
            break
    match_anio = re.search(r"(20\d{2})", nombre_archivo)
    anio = int(match_anio.group(1)) if match_anio else None
    fechas_validas = serie_fechas.dropna() if serie_fechas is not None else pd.Series(dtype="datetime64[ns]")
    if not fechas_validas.empty:
        if mes_num is None:
            mes_num = int(fechas_validas.dt.month.mode()[0])
        if anio is None:
            anio = int(fechas_validas.dt.year.mode()[0])
    if mes_num and anio:
        return f"{MESES_ES[mes_num][:3]} {anio}", anio * 100 + mes_num
    return nombre_archivo, 999999


# --------------------------------------------------------------
# LECTURA: CONSIGNACIONES
# --------------------------------------------------------------
@st.cache_data(show_spinner=False)
def leer_consignaciones(archivo_bytes, nombre_archivo):
    xls = pd.ExcelFile(archivo_bytes)
    hoja = detectar_hoja(xls, "consign")
    if hoja is None:
        return None, "no_encontrada", None
    try:
        df = pd.read_excel(xls, sheet_name=hoja, header=FILA_ENCABEZADO_CONSIGNACIONES, usecols=COLUMNAS_EXCEL_CONSIGNACIONES)
    except Exception as e:
        return None, "error_lectura", str(e)
    if len(df.columns) != len(COLUMNAS_INTERNAS_CONSIGNACIONES):
        return None, "columnas_incorrectas", None
    df.columns = COLUMNAS_INTERNAS_CONSIGNACIONES
    return df, hoja, None


def limpiar_consignaciones(df):
    df = df.copy()
    df["Fecha_Pago"] = pd.to_datetime(df["Fecha_Pago"], dayfirst=True, errors="coerce")
    for col in ["Valor_Mercancia", "Envio", "Consignacion_Neto"]:
        df[col] = df[col].apply(limpiar_valor_numerico)
    df["Medio_Pago"] = df["Medio_Pago"].astype(str).str.strip()
    df["Retribucion"] = df["Retribucion"].astype(str).str.strip()
    df["Descripcion"] = df["Descripcion"].astype(str).str.strip()
    for col in ["Medio_Pago", "Retribucion", "Descripcion"]:
        df.loc[df[col].isin(["nan", "None"]), col] = ""
    df = df.dropna(subset=["Fecha_Pago"])
    df["Medio_Pago"] = df["Medio_Pago"].replace("", "Sin especificar")
    for col in ["Valor_Mercancia", "Envio", "Consignacion_Neto"]:
        df[col] = df[col].fillna(0)
    return df


def explotar_productos(df):
    filas = []
    for _, fila in df.iterrows():
        productos = [p.strip() for p in str(fila["Descripcion"]).split(" - ") if p.strip()]
        for p in productos:
            filas.append({"Producto": p, "Periodo": fila["Periodo"], "Periodo_Orden": fila["Periodo_Orden"]})
    if not filas:
        return pd.DataFrame(columns=["Producto", "Periodo", "Periodo_Orden"])
    return pd.DataFrame(filas)


# --------------------------------------------------------------
# LECTURA: INVENTARIO
# --------------------------------------------------------------
@st.cache_data(show_spinner=False)
def leer_inventario(archivo_bytes, nombre_archivo):
    xls = pd.ExcelFile(archivo_bytes)
    hoja = detectar_hoja(xls, "inventario")
    if hoja is None:
        return None, "no_encontrada"
    try:
        df_bruto = pd.read_excel(xls, sheet_name=hoja, header=0)
    except Exception as e:
        return None, f"error_lectura: {e}"

    lookup = construir_lookup(df_bruto.columns, ALIAS_ENCABEZADOS_INVENTARIO)
    campos_requeridos = {
        "Producto": obtener_columna(lookup, "producto"),
        "Compra_Origen": obtener_columna(lookup, "compraorigen"),
        "Cantidad_Fisica": obtener_columna(lookup, "cantidadvendidafisica"),
        "Cantidad_Online": obtener_columna(lookup, "cantidadvendidaonline"),
        "Cantidad_Nomina": obtener_columna(lookup, "cantdeduccionnomina"),
        "Total_Vendido": obtener_columna(lookup, "totalvendido"),
        "Ingresos_Totales": obtener_columna(lookup, "ingresostotales"),
    }
    faltantes = [k for k, v in campos_requeridos.items() if v is None]
    if faltantes:
        return None, f"faltan_columnas: {', '.join(faltantes)}"

    # Opcionales: si faltan, no se cae la lectura — solo quedan en 0 (se usan para
    # rellenar la hoja Retribución FONDO, no son críticos para el resto de la app).
    campos_opcionales = {
        "Fecha": obtener_columna(lookup, "fecha"),
        "Inventario_Inicial": obtener_columna(lookup, "inventarioinicial"),
        "Costo_Unitario": obtener_columna(lookup, "costounitario"),
        "Costo_Total_Compra": obtener_columna(lookup, "costototalcompra"),
        "Precio_Venta_Unitario": obtener_columna(lookup, "precioventaunitario"),
        "Inventario_Final": obtener_columna(lookup, "inventariofinal"),
        "Utilidad_Bruta": obtener_columna(lookup, "utilidadbruta"),
    }

    campos = {**campos_requeridos, **campos_opcionales}
    df = pd.DataFrame()
    for nombre_interno, col_original in campos.items():
        df[nombre_interno] = df_bruto[col_original] if col_original is not None else pd.NA
    return df, None


def limpiar_inventario(df):
    df = df.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors="coerce")
    df["Producto"] = df["Producto"].astype(str).str.strip()
    df.loc[df["Producto"].isin(["nan", "None", ""]), "Producto"] = None
    df["Compra_Origen"] = df["Compra_Origen"].astype(str).str.strip()
    df.loc[df["Compra_Origen"].isin(["nan", "None"]), "Compra_Origen"] = ""
    campos_numericos = [
        "Cantidad_Fisica", "Cantidad_Online", "Cantidad_Nomina", "Total_Vendido", "Ingresos_Totales",
        "Inventario_Inicial", "Costo_Unitario", "Costo_Total_Compra", "Precio_Venta_Unitario",
        "Inventario_Final", "Utilidad_Bruta",
    ]
    for col in campos_numericos:
        df[col] = df[col].apply(limpiar_valor_numerico).fillna(0)
    df = df.dropna(subset=["Producto"])
    return df


def normalizar_nombre_fondo(nombre):
    """Igual que _normalizar_producto del script original: recorta una 'F' final suelta (ej. 'Camiseta M F')."""
    texto = re.sub(r"\s+", " ", str(nombre).strip())
    texto = re.sub(r"\s+f$", "", texto, flags=re.IGNORECASE)
    texto = "".join(c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn")
    return texto.strip().lower()


@st.cache_data(show_spinner=False)
def leer_master_fondo(archivo_bytes, nombre_archivo):
    """Lee la lista maestra de productos del Fondo desde la hoja 'Retribución FONDO' (columna B = Producto, D = Compra Origen)."""
    xls = pd.ExcelFile(archivo_bytes)
    hoja = detectar_hoja(xls, "retribuci")
    if hoja is None:
        return pd.DataFrame(columns=["Producto", "Compra_Origen"])
    try:
        grid = pd.read_excel(xls, sheet_name=hoja, header=None)
    except Exception:
        return pd.DataFrame(columns=["Producto", "Compra_Origen"])
    filas_master = []
    r = 1  # fila 2 de Excel = índice 1
    while r < grid.shape[0]:
        producto = grid.iat[r, 1] if grid.shape[1] > 1 else None
        if not isinstance(producto, str) or producto.strip() == "":
            break
        origen = grid.iat[r, 3] if grid.shape[1] > 3 else ""
        filas_master.append({"Producto": producto.strip(), "Compra_Origen": str(origen).strip() if pd.notna(origen) else ""})
        r += 1
    return pd.DataFrame(filas_master)


# --------------------------------------------------------------
# LECTURA: TASAS DE RETRIBUCIÓN AL FONDO
# --------------------------------------------------------------
@st.cache_data(show_spinner=False)
def leer_tasas_retribucion(archivo_bytes, nombre_archivo):
    xls = pd.ExcelFile(archivo_bytes)
    hoja = detectar_hoja(xls, "retribuci")
    if hoja is None:
        return {}
    try:
        grid = pd.read_excel(xls, sheet_name=hoja, header=None)
    except Exception:
        return {}
    filas, columnas = grid.shape
    fila_header, col_articulo, col_total = None, None, None
    for r in range(filas):
        for c in range(columnas):
            v = grid.iat[r, c]
            if isinstance(v, str):
                nv = normalizar_texto(v)
                if nv == "articulo":
                    fila_header, col_articulo = r, c
                if nv.startswith("totalretribucion"):
                    col_total = c
        if fila_header is not None and col_total is not None:
            break
    if fila_header is None or col_total is None:
        return {}
    tasas = {}
    r = fila_header + 1
    while r < filas:
        categoria = grid.iat[r, col_articulo]
        tasa = grid.iat[r, col_total]
        if not isinstance(categoria, str) or categoria.strip() == "":
            break
        tasas[normalizar_texto(categoria)] = float(tasa) if pd.notna(tasa) else None
        r += 1
    return tasas


def es_producto_fondo(compra_origen, producto):
    """Un producto es del Fondo si su Compra Origen lo indica, O si su nombre termina en
    'F' (ej. 'Camiseta Azul M F') — esta segunda señal es la que manda cuando el origen
    viene ambiguo/mixto (ej. 'UCN - FONDO' a nivel de pedido en Consignaciones)."""
    if isinstance(compra_origen, str) and "fondo" in compra_origen.lower():
        return True
    if isinstance(producto, str) and re.search(r"\s+F$", producto.strip(), flags=re.IGNORECASE):
        return True
    return False


def calcular_retribucion(df_inventario, tasas):
    mask_fondo = df_inventario.apply(lambda r: es_producto_fondo(r["Compra_Origen"], r["Producto"]), axis=1)
    df_fondo = df_inventario[mask_fondo & (df_inventario["Total_Vendido"] > 0)].copy()
    if df_fondo.empty:
        return df_fondo.assign(Categoria=[], Tasa=[], Retribucion=[])
    df_fondo["Categoria"] = df_fondo["Producto"].apply(categoria_retribucion)
    df_fondo["Tasa"] = df_fondo["Categoria"].apply(lambda c: tasas.get(normalizar_texto(c)) if c else None)
    df_fondo["Retribucion"] = df_fondo["Total_Vendido"] * df_fondo["Tasa"]
    return df_fondo


CAMPOS_HOJA_FONDO = [
    "Inventario_Inicial", "Costo_Unitario", "Costo_Total_Compra", "Precio_Venta_Unitario",
    "Cantidad_Fisica", "Cantidad_Online", "Cantidad_Nomina", "Total_Vendido",
    "Ingresos_Totales", "Inventario_Final", "Utilidad_Bruta",
]


def calcular_hoja_retribucion_fondo(df_inventario, master_fondo, tasas):
    """Reproduce la hoja 'Retribución FONDO' completa: una fila por producto de la lista
    maestra (aunque no haya vendido nada este periodo), con sus datos de Inventario y la
    retribución calculada — igual que actualizar_retribucion_fondo() del script original."""
    if master_fondo is None or master_fondo.empty or df_inventario is None or df_inventario.empty:
        return pd.DataFrame()

    mask_fondo = df_inventario.apply(lambda r: es_producto_fondo(r["Compra_Origen"], r["Producto"]), axis=1)
    df_fondo = df_inventario[mask_fondo].copy()
    df_fondo["Clave"] = df_fondo["Producto"].apply(normalizar_nombre_fondo)

    agregados = {"Costo_Unitario": "first", "Precio_Venta_Unitario": "first"}
    for c in CAMPOS_HOJA_FONDO:
        agregados.setdefault(c, "sum")
    agregado = df_fondo.groupby("Clave", as_index=False).agg(agregados)
    lookup = agregado.set_index("Clave").to_dict("index")

    filas = []
    for _, fila_master in master_fondo.iterrows():
        nombre = fila_master["Producto"]
        clave = normalizar_nombre_fondo(nombre)
        datos = lookup.get(clave)
        categoria = categoria_retribucion(nombre)
        tasa = tasas.get(normalizar_texto(categoria)) if categoria else None

        fila = {"Producto": nombre, "Compra_Origen": fila_master["Compra_Origen"], "Categoria": categoria, "Tasa": tasa}
        for c in CAMPOS_HOJA_FONDO:
            fila[c] = float(datos[c]) if datos is not None else 0.0
        total_vendido = fila["Total_Vendido"]
        fila["Retribucion"] = (total_vendido * tasa) if (tasa is not None and total_vendido) else 0.0
        fila["Sin_Tasa"] = tasa is None and total_vendido > 0
        filas.append(fila)

    return pd.DataFrame(filas)


# --------------------------------------------------------------
# CONTROL DE CALIDAD: CONSIGNACIONES vs INVENTARIO
# --------------------------------------------------------------
def comparar_consignaciones_inventario(productos_pedidos_df, df_inventario):
    """Para cada producto individual de un pedido, busca coincidencia parcial en Inventario."""
    productos_inv_norm = set(normalizar_texto(p) for p in df_inventario["Producto"].dropna() if str(p).strip())
    resultado = productos_pedidos_df.copy()
    if resultado.empty:
        resultado["Coincide"] = []
        return resultado

    def coincide(producto):
        pn = normalizar_texto(producto)
        if not pn:
            return True
        return any(pn in inv or inv in pn for inv in productos_inv_norm if inv)

    resultado["Coincide"] = resultado["Producto"].apply(coincide)
    return resultado


# --------------------------------------------------------------
# EXPORTACIÓN: EXCEL DE ANÁLISIS DE FIN DE MES
# --------------------------------------------------------------
def _hs(cell, color=XL_COLOR_HEADER):
    cell.font = Font(name=XL_FUENTE, bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", start_color=color, end_color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _borde(cell)


def _ss(cell):
    cell.font = Font(name=XL_FUENTE, bold=True, size=10)
    cell.fill = PatternFill("solid", start_color=XL_COLOR_SUBTOTAL, end_color=XL_COLOR_SUBTOTAL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    _borde(cell)


def _ds(cell, bg=None, bold=False):
    cell.font = Font(name=XL_FUENTE, bold=bold, size=10)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _borde(cell)


def _borde(cell):
    thin = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _escribir_hoja_analisis(ws, periodo, datos, resumen_ia=None):
    """Llena una hoja (ya creada) con el contenido del análisis de un periodo."""
    ws.sheet_view.showGridLines = False
    fila = 1

    ws.merge_cells(f"A{fila}:F{fila}")
    c = ws[f"A{fila}"]
    c.value = f"ANÁLISIS DE VENTAS — {periodo} — Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = Font(name=XL_FUENTE, bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="1A3A6B", end_color="1A3A6B")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 26
    fila += 2

    # 1. Recuento de prendas vendidas
    ws.merge_cells(f"A{fila}:F{fila}")
    _hs(ws[f"A{fila}"])
    ws[f"A{fila}"].value = "1. RECUENTO DE PRENDAS VENDIDAS (Inventario)"
    fila += 1
    for j, h in enumerate(["Canal", "Cantidad"], start=1):
        _hs(ws.cell(row=fila, column=j, value=h))
    fila += 1
    df_inv = datos["inventario"]
    canales = [
        ("Física", df_inv["Cantidad_Fisica"].sum()),
        ("Online", df_inv["Cantidad_Online"].sum()),
        ("Deducción Nómina", df_inv["Cantidad_Nomina"].sum()),
    ]
    for label, val in canales:
        _ds(ws.cell(row=fila, column=1, value=label))
        _ds(ws.cell(row=fila, column=2, value=int(val)))
        fila += 1
    total_prendas = sum(v for _, v in canales)
    _ss(ws.cell(row=fila, column=1, value="TOTAL PRENDAS VENDIDAS"))
    _ss(ws.cell(row=fila, column=2, value=int(total_prendas)))
    fila += 2

    # 2. Comparación Consignaciones vs Inventario
    ws.merge_cells(f"A{fila}:F{fila}")
    _hs(ws[f"A{fila}"])
    ws[f"A{fila}"].value = "2. CONTROL DE CALIDAD: PEDIDOS vs INVENTARIO"
    fila += 1
    for j, h in enumerate(["Producto (pedido)", "¿En Inventario?"], start=1):
        _hs(ws.cell(row=fila, column=j, value=h))
    fila += 1
    comparacion = datos["comparacion"]
    if comparacion.empty:
        _ds(ws.cell(row=fila, column=1, value="Sin pedidos para este periodo"))
        fila += 1
    else:
        for _, r in comparacion.iterrows():
            bg = XL_COLOR_OK if r["Coincide"] else XL_COLOR_ALERT
            _ds(ws.cell(row=fila, column=1, value=r["Producto"]), bg=bg)
            _ds(ws.cell(row=fila, column=2, value="SÍ" if r["Coincide"] else "NO ENCONTRADO"), bg=bg)
            fila += 1
    fila += 1

    # 3. Ventas netas
    ws.merge_cells(f"A{fila}:F{fila}")
    _hs(ws[f"A{fila}"])
    ws[f"A{fila}"].value = "3. TOTAL VENTAS NETAS"
    fila += 1
    for j, h in enumerate(["Fuente", "Monto ($)"], start=1):
        _hs(ws.cell(row=fila, column=j, value=h))
    fila += 1
    df_cons = datos["consignaciones"]
    filas_ventas = [
        ("Ingresos Totales (Inventario)", df_inv["Ingresos_Totales"].sum()),
        ("Valor Mercancía (Consignaciones, bruto)", df_cons["Valor_Mercancia"].sum()),
        ("Consignación recibida (neto)", df_cons["Consignacion_Neto"].sum()),
    ]
    for label, val in filas_ventas:
        _ds(ws.cell(row=fila, column=1, value=label))
        c2 = ws.cell(row=fila, column=2, value=float(val))
        _ds(c2)
        c2.number_format = "$#,##0"
        fila += 1
    fila += 1

    # 4. Retribución al Fondo de Empleados
    ws.merge_cells(f"A{fila}:F{fila}")
    _hs(ws[f"A{fila}"])
    ws[f"A{fila}"].value = "4. RETRIBUCIÓN AL FONDO DE EMPLEADOS"
    fila += 1
    encabezados_r = ["Producto", "Categoría", "Unidades Vendidas", "Tasa por Unidad ($)", "Retribución ($)"]
    for j, h in enumerate(encabezados_r, start=1):
        _hs(ws.cell(row=fila, column=j, value=h))
    fila += 1
    retribucion = datos["retribucion"]
    total_retribucion = 0
    if retribucion.empty:
        _ds(ws.cell(row=fila, column=1, value="No hay productos del Fondo con ventas este periodo"))
        fila += 1
    else:
        for _, r in retribucion.iterrows():
            sin_tasa = pd.isna(r["Tasa"])
            bg = XL_COLOR_ALERT if sin_tasa else None
            _ds(ws.cell(row=fila, column=1, value=r["Producto"]), bg=bg)
            _ds(ws.cell(row=fila, column=2, value=r["Categoria"] or "Sin categorizar"), bg=bg)
            _ds(ws.cell(row=fila, column=3, value=int(r["Total_Vendido"])), bg=bg)
            c_tasa = ws.cell(row=fila, column=4, value=None if sin_tasa else float(r["Tasa"]))
            c_retrib = ws.cell(row=fila, column=5, value=None if sin_tasa else float(r["Retribucion"]))
            _ds(c_tasa, bg=bg)
            _ds(c_retrib, bg=bg)
            if not sin_tasa:
                c_tasa.number_format = "$#,##0"
                c_retrib.number_format = "$#,##0"
                total_retribucion += r["Retribucion"]
            fila += 1
    _ss(ws.cell(row=fila, column=1, value="TOTAL A RETRIBUIR"))
    ws.merge_cells(f"A{fila}:D{fila}")
    c_total = ws.cell(row=fila, column=5, value=float(total_retribucion))
    _ss(c_total)
    c_total.number_format = "$#,##0"
    fila += 2

    # 5. Resumen generado por IA (opcional)
    if resumen_ia:
        ws.merge_cells(f"A{fila}:F{fila}")
        _hs(ws[f"A{fila}"])
        ws[f"A{fila}"].value = "5. RESUMEN GENERADO POR IA"
        fila += 1
        filas_texto = max(6, len(resumen_ia) // 80 + 2)
        ws.merge_cells(f"A{fila}:F{fila + filas_texto - 1}")
        celda = ws.cell(row=fila, column=1, value=resumen_ia)
        celda.font = Font(name=XL_FUENTE, size=10)
        celda.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        celda.fill = PatternFill("solid", start_color=XL_COLOR_SUBTOTAL, end_color=XL_COLOR_SUBTOTAL)
        ws.row_dimensions[fila].height = 15 * filas_texto
        fila += filas_texto

    widths = {"A": 32, "B": 22, "C": 16, "D": 16, "E": 16, "F": 16}
    for col_l, w in widths.items():
        ws.column_dimensions[col_l].width = w


# Columnas reales de la hoja "Retribución FONDO" en el archivo original (igual que Inventario)
COLUMNAS_EXCEL_HOJA_FONDO = {
    "Inventario_Inicial": 5, "Costo_Unitario": 6, "Costo_Total_Compra": 7, "Precio_Venta_Unitario": 8,
    "Cantidad_Fisica": 9, "Cantidad_Online": 10, "Cantidad_Nomina": 11, "Total_Vendido": 12,
    "Ingresos_Totales": 13, "Inventario_Final": 14, "Utilidad_Bruta": 15, "Retribucion": 16,
}
COLUMNAS_MONEDA_HOJA_FONDO = [6, 7, 8, 13, 15, 16]  # Costo Unitario, Costo Total, Precio Venta, Ingresos, Utilidad, Retribución


def _actualizar_hoja_fondo_original(wb, hoja_fondo):
    """Rellena, DENTRO del propio archivo original, las columnas E-P de la hoja real
    'Retribución FONDO' (fila 2 en adelante = lista maestra), igual que hace el script."""
    if hoja_fondo is None or hoja_fondo.empty:
        return False
    nombre_hoja = detectar_hoja(_FakeXls(wb.sheetnames), "retribuci")
    if nombre_hoja is None:
        return False
    ws = wb[nombre_hoja]

    num_productos = len(hoja_fondo)
    ultima_fila = 1 + num_productos
    for fila_excel in range(2, ultima_fila + 1):
        for col in range(5, 17):
            ws.cell(row=fila_excel, column=col).value = None

    for i, (_, r) in enumerate(hoja_fondo.iterrows()):
        fila_excel = 2 + i
        for campo, col in COLUMNAS_EXCEL_HOJA_FONDO.items():
            valor = r["Retribucion"] if campo == "Retribucion" else r[campo]
            celda = ws.cell(row=fila_excel, column=col, value=float(valor))
            if col in COLUMNAS_MONEDA_HOJA_FONDO:
                celda.number_format = "$#,##0"
        if r.get("Sin_Tasa"):
            relleno = PatternFill("solid", start_color=XL_COLOR_ALERT, end_color=XL_COLOR_ALERT)
            for col in range(2, 17):
                ws.cell(row=fila_excel, column=col).fill = relleno

    fila_total = ultima_fila + 1
    ws.cell(row=fila_total, column=14, value="TOTAL A RETRIBUIR:").font = Font(name=XL_FUENTE, bold=True)
    c_total = ws.cell(row=fila_total, column=16, value=float(hoja_fondo["Retribucion"].sum()))
    c_total.font = Font(name=XL_FUENTE, bold=True)
    c_total.number_format = "$#,##0"
    return True


class _FakeXls:
    """Envoltorio mínimo para reutilizar detectar_hoja() sobre una lista de nombres de hoja."""
    def __init__(self, sheet_names):
        self.sheet_names = sheet_names


def generar_workbook_completo(archivo_bytes, periodo, datos, resumen_ia=None):
    """Abre el archivo ORIGINAL del usuario y le agrega/actualiza ahí mismo la hoja de
    Análisis y la hoja real 'Retribución FONDO' — un solo archivo completo de salida."""
    wb = load_workbook(io.BytesIO(archivo_bytes))

    nombre_analisis = f"Analisis {periodo}"[:31]
    if nombre_analisis in wb.sheetnames:
        del wb[nombre_analisis]
    ws = wb.create_sheet(nombre_analisis)
    _escribir_hoja_analisis(ws, periodo, datos, resumen_ia)

    hoja_fondo = datos.get("hoja_fondo")
    _actualizar_hoja_fondo_original(wb, hoja_fondo)

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida


# --------------------------------------------------------------
# ASISTENTE DE IA (Groq)
# --------------------------------------------------------------
def obtener_api_key_groq():
    """Busca la key primero en st.secrets (uso permanente) y si no existe, deja que el usuario la pegue en la sesión."""
    try:
        clave_guardada = st.secrets.get("GROQ_API_KEY", "")
    except Exception:
        clave_guardada = ""
    if clave_guardada:
        return clave_guardada
    return st.session_state.get("groq_api_key_sesion", "")


def construir_contexto_periodo(periodo, datos, incluir_pedidos=True):
    """Arma un resumen de texto compacto de un periodo para pasarle a la IA como contexto."""
    df_inv = datos["inventario"]
    df_cons = datos["consignaciones"]
    productos_pedidos = datos["productos_pedidos"]
    comparacion = datos["comparacion"]
    retribucion = datos["retribucion"]

    partes = [f"=== DATOS DEL PERIODO: {periodo} ==="]

    ingresos_neto = df_cons["Consignacion_Neto"].sum()
    valor_bruto = df_cons["Valor_Mercancia"].sum()
    partes.append(f"Ingresos netos (Consignación): {formato_pesos(ingresos_neto)}")
    partes.append(f"Valor Mercancía (bruto): {formato_pesos(valor_bruto)}")
    partes.append(f"Pedidos totales: {len(df_cons)}")
    partes.append(f"Unidades vendidas (según pedidos): {len(productos_pedidos)}")

    if not df_cons.empty:
        por_canal = df_cons.groupby("Medio_Pago")["Consignacion_Neto"].sum().sort_values(ascending=False)
        partes.append("Ingresos por canal: " + ", ".join(f"{c}: {formato_pesos(v)}" for c, v in por_canal.items()))

    if not productos_pedidos.empty:
        top5 = productos_pedidos.groupby("Producto").size().sort_values(ascending=False).head(5)
        partes.append("Top 5 productos más vendidos: " + ", ".join(f"{p} ({n} uds)" for p, n in top5.items()))

    if not df_inv.empty:
        partes.append(
            f"Unidades por canal (Inventario) — Física: {int(df_inv['Cantidad_Fisica'].sum())}, "
            f"Online: {int(df_inv['Cantidad_Online'].sum())}, Nómina: {int(df_inv['Cantidad_Nomina'].sum())}"
        )
        partes.append(f"Ingresos Totales (Inventario): {formato_pesos(df_inv['Ingresos_Totales'].sum())}")

    if not comparacion.empty:
        sin_match = comparacion[~comparacion["Coincide"]]
        if sin_match.empty:
            partes.append("Control de calidad: todos los productos de los pedidos coinciden con Inventario.")
        else:
            partes.append(
                f"Control de calidad: {len(sin_match)} producto(s) de pedidos sin coincidencia en Inventario: "
                + ", ".join(sin_match['Producto'].unique()[:10])
            )

    if not retribucion.empty:
        total_r = retribucion["Retribucion"].sum(skipna=True)
        partes.append(f"Retribución al Fondo de Empleados: {formato_pesos(total_r)}")
        detalle_r = "; ".join(
            f"{r['Producto']} ({r['Categoria'] or 'sin categorizar'}): {int(r['Total_Vendido'])} uds"
            for _, r in retribucion.iterrows()
        )
        partes.append(f"Detalle retribución: {detalle_r}")

    if incluir_pedidos and not df_cons.empty:
        cols = ["Fecha_Pago", "Medio_Pago", "Descripcion", "Valor_Mercancia", "Consignacion_Neto"]
        muestra = df_cons[cols].head(200).copy()
        muestra["Fecha_Pago"] = muestra["Fecha_Pago"].dt.strftime("%d/%m/%Y")
        partes.append("Detalle de pedidos (fecha, canal, descripción, valor mercancía, consignación neta):")
        partes.append(muestra.to_csv(index=False))

    return "\n".join(partes)


def preguntar_ia(pregunta, contexto, api_key, historial=None):
    cliente = Groq(api_key=api_key)
    mensajes = [{
        "role": "system",
        "content": (
            "Eres un analista de datos para la Tienda Virtual UCN. Respondes preguntas sobre las "
            "ventas, inventario y retribución al Fondo de Empleados, basándote únicamente en los datos "
            "que se te entregan. Si algo no está en los datos, dilo claramente en vez de inventar. "
            "Responde en español, de forma breve y directa."
        ),
    }]
    if historial:
        mensajes.extend(historial)
    mensajes.append({"role": "user", "content": f"DATOS:\n{contexto}\n\nPREGUNTA: {pregunta}"})
    respuesta = cliente.chat.completions.create(model=MODELO_GROQ, messages=mensajes, temperature=0.3, max_tokens=800)
    return respuesta.choices[0].message.content


def generar_resumen_ia(periodo, contexto, api_key):
    cliente = Groq(api_key=api_key)
    mensajes = [
        {
            "role": "system",
            "content": (
                "Eres un analista de negocio para la Tienda Virtual UCN. Escribe un resumen ejecutivo "
                "breve (3-5 párrafos cortos) del desempeño del mes con base en los datos entregados: "
                "qué se vendió, qué canal predominó, hallazgos del control de calidad y de la retribución "
                "al Fondo, y 2-3 recomendaciones concretas. Español, tono profesional y directo, sin inventar cifras."
            ),
        },
        {"role": "user", "content": f"DATOS DEL PERIODO {periodo}:\n{contexto}"},
    ]
    respuesta = cliente.chat.completions.create(model=MODELO_GROQ, messages=mensajes, temperature=0.4, max_tokens=900)
    return respuesta.choices[0].message.content


# --------------------------------------------------------------
# ESTADO / TEMA
# --------------------------------------------------------------
if "tema" not in st.session_state:
    st.session_state["tema"] = "Claro"

st.sidebar.markdown(
    """<div class="sidebar-brand"><div class="name">TIENDA VIRTUAL UCN</div>
    <div class="tag">Reporte de ventas</div></div>""",
    unsafe_allow_html=True,
)

st.sidebar.markdown('<div class="sidebar-section">Apariencia</div>', unsafe_allow_html=True)
st.session_state["tema"] = st.sidebar.radio(
    "Tema", options=["Claro", "Oscuro"], horizontal=True,
    index=["Claro", "Oscuro"].index(st.session_state["tema"]), label_visibility="collapsed",
)
tema = TEMAS[st.session_state["tema"]]
inyectar_estilos(tema)

st.sidebar.markdown('<div class="sidebar-section">Archivos mensuales</div>', unsafe_allow_html=True)
archivos_subidos = st.sidebar.file_uploader(
    "Cargar uno o varios archivos (.xlsx)", type=["xlsx"], accept_multiple_files=True,
    label_visibility="collapsed",
    help="Puedes cargar varios archivos mensuales a la vez (ej: Junio, Julio, Agosto) para compararlos.",
)

st.markdown(
    """<div class="app-header"><div class="eyebrow">Tienda Virtual UCN</div>
    <div class="title">Reporte de Ventas</div>
    <div class="subtitle">Ingresos, canal de venta, control de calidad y retribución al Fondo — todo desde tus archivos mensuales.</div></div>""",
    unsafe_allow_html=True,
)

if not archivos_subidos:
    st.info("Carga uno o varios archivos \"Control_Integral_Inventario_<MES>.xlsx\" desde el panel izquierdo.")
    st.caption(
        "La app lee la hoja 'Consignaciones' (pedidos), la hoja 'Inventario' (catálogo y ventas por canal) "
        "y la hoja 'Retribución FONDO' (tasas) — ese formato es fijo."
    )
    st.stop()

# --------------------------------------------------------------
# LECTURA DE TODOS LOS ARCHIVOS
# --------------------------------------------------------------
info_archivos = []
for archivo in archivos_subidos:
    df_cons_bruto, hoja_cons, detalle_error = leer_consignaciones(archivo, archivo.name)
    if df_cons_bruto is None:
        if hoja_cons == "no_encontrada":
            st.error(f"{archivo.name}: no se encontró una hoja llamada 'Consignaciones'.")
        elif hoja_cons == "columnas_incorrectas":
            st.error(f"{archivo.name}: la hoja 'Consignaciones' no tiene las columnas esperadas (F, G, I, J, N, O, P).")
        else:
            st.error(f"{archivo.name}: no se pudo leer 'Consignaciones' ({detalle_error}).")
        continue
    df_cons = limpiar_consignaciones(df_cons_bruto)

    df_inv_bruto, error_inv = leer_inventario(archivo, archivo.name)
    df_inv = limpiar_inventario(df_inv_bruto) if df_inv_bruto is not None else None
    if df_inv is None:
        st.warning(
            f"{archivo.name}: no se pudo leer la hoja 'Inventario' correctamente ({error_inv}). "
            "El control de calidad y la retribución al Fondo no estarán disponibles para este archivo."
        )

    tasas = leer_tasas_retribucion(archivo, archivo.name)
    if df_inv is not None and not tasas:
        st.warning(f"{archivo.name}: no se encontró la tabla de tasas en 'Retribución FONDO'; la retribución no se calculará para este archivo.")

    master_fondo = leer_master_fondo(archivo, archivo.name)
    if df_inv is not None and master_fondo.empty:
        st.warning(f"{archivo.name}: no se encontró la lista de productos en 'Retribución FONDO'; esa hoja no se generará en el Excel para este archivo.")

    etiqueta_sugerida, orden_sugerido = detectar_periodo(archivo.name, df_cons["Fecha_Pago"])
    info_archivos.append({
        "archivo": archivo.name, "archivo_bytes": archivo.getvalue(), "df_cons": df_cons, "df_inv": df_inv,
        "tasas": tasas, "master_fondo": master_fondo,
        "etiqueta_sugerida": etiqueta_sugerida, "orden_sugerido": orden_sugerido,
    })

if not info_archivos:
    st.stop()

todas_confiables = all(info["orden_sugerido"] != 999999 for info in info_archivos)
with st.expander("Periodos detectados por archivo", expanded=not todas_confiables):
    for info in info_archivos:
        info["etiqueta_final"] = st.text_input(
            info["archivo"], value=info["etiqueta_sugerida"], key=f"periodo_{info['archivo']}",
        ) or info["etiqueta_sugerida"]

partes_cons = []
datos_por_periodo = {}
for info in info_archivos:
    periodo = info["etiqueta_final"]
    df_cons_parte = info["df_cons"].copy()
    df_cons_parte["Periodo"] = periodo
    df_cons_parte["Periodo_Orden"] = info["orden_sugerido"]
    df_cons_parte["Archivo"] = info["archivo"]
    partes_cons.append(df_cons_parte)

    productos_pedidos = explotar_productos(df_cons_parte)
    comparacion = comparar_consignaciones_inventario(productos_pedidos, info["df_inv"]) if info["df_inv"] is not None else pd.DataFrame(columns=["Producto", "Periodo", "Periodo_Orden", "Coincide"])
    retribucion = calcular_retribucion(info["df_inv"], info["tasas"]) if info["df_inv"] is not None else pd.DataFrame()
    hoja_fondo = calcular_hoja_retribucion_fondo(info["df_inv"], info["master_fondo"], info["tasas"]) if info["df_inv"] is not None else pd.DataFrame()

    datos_por_periodo[periodo] = {
        "archivo_bytes": info["archivo_bytes"], "archivo_nombre": info["archivo"],
        "inventario": info["df_inv"] if info["df_inv"] is not None else pd.DataFrame(
            columns=["Fecha", "Producto", "Compra_Origen", "Cantidad_Fisica", "Cantidad_Online", "Cantidad_Nomina", "Total_Vendido", "Ingresos_Totales"]
        ),
        "consignaciones": df_cons_parte,
        "productos_pedidos": productos_pedidos,
        "comparacion": comparacion,
        "retribucion": retribucion,
        "hoja_fondo": hoja_fondo,
    }

df = pd.concat(partes_cons, ignore_index=True)
if df.empty:
    st.error("No quedaron pedidos válidos (con fecha de pago) en los archivos cargados.")
    st.stop()

orden_periodos = df[["Periodo", "Periodo_Orden"]].drop_duplicates().sort_values("Periodo_Orden")["Periodo"].tolist()

# --------------------------------------------------------------
# FILTROS
# --------------------------------------------------------------
st.sidebar.markdown('<div class="sidebar-section">Filtros</div>', unsafe_allow_html=True)
periodos_seleccionados = st.sidebar.multiselect("Periodo", options=orden_periodos, default=orden_periodos)

canales_disponibles = sorted(df["Medio_Pago"].unique().tolist())
canales_seleccionados = st.sidebar.multiselect("Canal de venta (Medio de Pago)", options=canales_disponibles, default=canales_disponibles)

df_filtrado = df[df["Periodo"].isin(periodos_seleccionados) & df["Medio_Pago"].isin(canales_seleccionados)]

st.sidebar.markdown('<div class="sidebar-section">Resumen de carga</div>', unsafe_allow_html=True)
for info in info_archivos:
    st.sidebar.caption(f"{info['etiqueta_final']} — {info['archivo']} ({len(info['df_cons'])} pedidos)")

st.sidebar.markdown('<div class="sidebar-section">Asistente de IA (Groq)</div>', unsafe_allow_html=True)
if not GROQ_DISPONIBLE:
    st.sidebar.caption("Falta instalar el paquete 'groq' (agregado a requirements.txt).")
groq_api_key = obtener_api_key_groq()
if GROQ_DISPONIBLE and not groq_api_key:
    clave_ingresada = st.sidebar.text_input(
        "Pega tu Groq API Key", type="password",
        help="Solo se usa en esta sesión, no se guarda en ningún archivo.",
    )
    if clave_ingresada:
        st.session_state["groq_api_key_sesion"] = clave_ingresada
        groq_api_key = clave_ingresada
        st.rerun()
elif GROQ_DISPONIBLE:
    st.sidebar.caption("API key cargada para esta sesión.")

st.sidebar.markdown('<div class="sidebar-section">Análisis de fin de mes</div>', unsafe_allow_html=True)
incluir_resumen_ia = False
if GROQ_DISPONIBLE and groq_api_key:
    incluir_resumen_ia = st.sidebar.checkbox("Incluir resumen ejecutivo generado por IA", value=False)

if st.sidebar.button("Generar Análisis de Fin de Mes", use_container_width=True):
    resumen_ia_por_periodo = {}
    if incluir_resumen_ia:
        with st.spinner("Generando resumen con IA..."):
            for periodo in periodos_seleccionados:
                try:
                    contexto = construir_contexto_periodo(periodo, datos_por_periodo[periodo], incluir_pedidos=False)
                    resumen_ia_por_periodo[periodo] = generar_resumen_ia(periodo, contexto, groq_api_key)
                except Exception as e:
                    st.sidebar.warning(f"No se pudo generar el resumen IA de {periodo}: {e}")

    with st.spinner("Generando archivo completo..."):
        if len(periodos_seleccionados) == 1:
            periodo = periodos_seleccionados[0]
            datos = datos_por_periodo[periodo]
            salida = generar_workbook_completo(datos["archivo_bytes"], periodo, datos, resumen_ia_por_periodo.get(periodo))
            nombre_base = datos["archivo_nombre"].rsplit(".", 1)[0]
            st.session_state["excel_analisis"] = salida
            st.session_state["excel_analisis_nombre"] = f"{nombre_base}_Actualizado.xlsx"
            st.session_state["excel_analisis_mime"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            buffer_zip = io.BytesIO()
            with zipfile.ZipFile(buffer_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                for periodo in periodos_seleccionados:
                    datos = datos_por_periodo[periodo]
                    salida = generar_workbook_completo(datos["archivo_bytes"], periodo, datos, resumen_ia_por_periodo.get(periodo))
                    nombre_base = datos["archivo_nombre"].rsplit(".", 1)[0]
                    zf.writestr(f"{nombre_base}_Actualizado.xlsx", salida.getvalue())
            buffer_zip.seek(0)
            st.session_state["excel_analisis"] = buffer_zip
            st.session_state["excel_analisis_nombre"] = f"Analisis_Ventas_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
            st.session_state["excel_analisis_mime"] = "application/zip"

if "excel_analisis" in st.session_state:
    if len(periodos_seleccionados) > 1:
        st.sidebar.caption("Como hay varios periodos, se genera un .zip con un Excel completo por cada mes.")
    st.sidebar.download_button(
        "Descargar Excel",
        data=st.session_state["excel_analisis"],
        file_name=st.session_state["excel_analisis_nombre"],
        mime=st.session_state["excel_analisis_mime"],
        use_container_width=True,
    )

if df_filtrado.empty:
    st.warning("No hay datos para los filtros seleccionados.")
    st.stop()

productos_filtrados = explotar_productos(df_filtrado)

# --------------------------------------------------------------
# KPIs
# --------------------------------------------------------------
ingresos_netos = df_filtrado["Consignacion_Neto"].sum()
valor_bruto = df_filtrado["Valor_Mercancia"].sum()
unidades_totales = len(productos_filtrados)
pedidos_totales = len(df_filtrado)

col1, col2, col3, col4 = st.columns(4)
render_kpi(col1, "Ingresos (Consignación neta)", formato_pesos(ingresos_netos), ACENTO_INGRESOS)
render_kpi(col2, "Valor Mercancía (bruto)", formato_pesos(valor_bruto), ACENTO_BRUTO)
render_kpi(col3, "Unidades vendidas", f"{unidades_totales:,.0f}".replace(",", "."), ACENTO_UNIDADES)
render_kpi(col4, "Pedidos totales", f"{pedidos_totales:,.0f}".replace(",", "."), ACENTO_PEDIDOS)

st.write("")
hay_comparacion = len(periodos_seleccionados) > 1

# --------------------------------------------------------------
# GRÁFICOS
# --------------------------------------------------------------
col_izq, col_der = st.columns(2)

with col_izq:
    st.markdown('<div class="section-title">Ingresos por Canal de Venta</div>', unsafe_allow_html=True)
    por_canal = df_filtrado.groupby("Medio_Pago", as_index=False)["Consignacion_Neto"].sum().sort_values("Consignacion_Neto", ascending=False)
    colores_canal = [PALETA_CATEGORIAS[i % len(PALETA_CATEGORIAS)] for i in range(len(por_canal))]
    fig_canal = px.bar(
        por_canal, x="Medio_Pago", y="Consignacion_Neto", template=tema["plotly_template"],
        text_auto=".2s", labels={"Medio_Pago": "", "Consignacion_Neto": "Ingresos ($)"},
    )
    fig_canal.update_traces(marker_color=colores_canal)
    fig_canal.update_layout(
        showlegend=False, paper_bgcolor=tema["chart_bg"], plot_bgcolor=tema["chart_bg"],
        font=dict(family="Inter, sans-serif", color=tema["texto"]), margin=dict(t=10, l=10, r=10, b=10),
    )
    st.plotly_chart(fig_canal, use_container_width=True)

with col_der:
    st.markdown('<div class="section-title">Ingresos vs. Valor Mercancía por Periodo</div>', unsafe_allow_html=True)
    resumen_periodo = (
        df_filtrado.groupby(["Periodo", "Periodo_Orden"], as_index=False)
        .agg(Consignacion_Neto=("Consignacion_Neto", "sum"), Valor_Mercancia=("Valor_Mercancia", "sum"))
        .sort_values("Periodo_Orden")
    )
    fig_periodo = go.Figure()
    fig_periodo.add_trace(go.Bar(
        x=resumen_periodo["Periodo"], y=resumen_periodo["Consignacion_Neto"],
        name="Ingresos (neto)", marker_color=ACENTO_INGRESOS,
        text=resumen_periodo["Consignacion_Neto"].apply(formato_pesos), textposition="outside",
    ))
    fig_periodo.add_trace(go.Bar(
        x=resumen_periodo["Periodo"], y=resumen_periodo["Valor_Mercancia"],
        name="Valor Mercancía (bruto)", marker_color=ACENTO_BRUTO,
        text=resumen_periodo["Valor_Mercancia"].apply(formato_pesos), textposition="outside",
    ))
    fig_periodo.update_layout(
        barmode="group", template=tema["plotly_template"],
        paper_bgcolor=tema["chart_bg"], plot_bgcolor=tema["chart_bg"],
        font=dict(family="Inter, sans-serif", color=tema["texto"]),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
        margin=dict(t=40, l=10, r=10, b=10),
    )
    st.plotly_chart(fig_periodo, use_container_width=True)

st.markdown('<div class="section-title">Top 5 Productos Más Vendidos (unidades)</div>', unsafe_allow_html=True)
if productos_filtrados.empty:
    st.caption("No hay productos identificables en las descripciones filtradas.")
else:
    top_total = productos_filtrados.groupby("Producto", as_index=False).size().rename(columns={"size": "Unidades"})
    top_total = top_total.sort_values("Unidades", ascending=False).head(5)
    productos_top = top_total["Producto"].tolist()
    data_top = productos_filtrados[productos_filtrados["Producto"].isin(productos_top)]

    if hay_comparacion:
        top_por_periodo = data_top.groupby(["Producto", "Periodo", "Periodo_Orden"], as_index=False).size().rename(columns={"size": "Unidades"})
        orden_categoria = top_total.sort_values("Unidades", ascending=True)["Producto"].tolist()
        fig_top = px.bar(
            top_por_periodo, x="Unidades", y="Producto", color="Periodo", orientation="h", barmode="group",
            template=tema["plotly_template"], category_orders={"Producto": orden_categoria, "Periodo": orden_periodos},
            color_discrete_sequence=PALETA_CATEGORIAS,
        )
    else:
        top_ordenado = top_total.sort_values("Unidades", ascending=True)
        fig_top = px.bar(top_ordenado, x="Unidades", y="Producto", orientation="h", template=tema["plotly_template"], text_auto=True)
        fig_top.update_traces(marker_color=ACENTO_INGRESOS)

    fig_top.update_layout(
        paper_bgcolor=tema["chart_bg"], plot_bgcolor=tema["chart_bg"],
        font=dict(family="Inter, sans-serif", color=tema["texto"]), margin=dict(t=10, l=10, r=10, b=10),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
        yaxis_title="", xaxis_title="Unidades vendidas",
    )
    st.plotly_chart(fig_top, use_container_width=True)

# --------------------------------------------------------------
# CONTROL DE CALIDAD: CONSIGNACIONES vs INVENTARIO
# --------------------------------------------------------------
st.write("")
st.markdown('<div class="section-title">Control de Calidad: Pedidos vs. Inventario</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="section-caption">Compara cada producto de los pedidos filtrados contra el catálogo de Inventario del mismo periodo.</div>',
    unsafe_allow_html=True,
)

comparaciones_filtradas = []
for periodo in periodos_seleccionados:
    comp = datos_por_periodo[periodo]["comparacion"]
    if not comp.empty:
        comparaciones_filtradas.append(comp)

if not comparaciones_filtradas:
    st.caption("No hay datos de Inventario disponibles para calcular el control de calidad.")
else:
    comp_total = pd.concat(comparaciones_filtradas, ignore_index=True)
    total_prod = len(comp_total)
    sin_match = comp_total[~comp_total["Coincide"]]
    n_sin_match = len(sin_match)

    if n_sin_match == 0:
        st.markdown(f'<div class="qa-ok">Los {total_prod} productos de los pedidos coinciden con el Inventario.</div>', unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div class="qa-alert">{n_sin_match} de {total_prod} productos de los pedidos no se encontraron en el Inventario del mismo periodo.</div>',
            unsafe_allow_html=True,
        )
        with st.expander("Ver productos sin coincidencia"):
            resumen_sin_match = sin_match.groupby(["Producto", "Periodo"], as_index=False).size().rename(columns={"size": "Veces"})
            st.dataframe(resumen_sin_match.sort_values("Veces", ascending=False), use_container_width=True, hide_index=True)

# --------------------------------------------------------------
# RETRIBUCIÓN AL FONDO DE EMPLEADOS
# --------------------------------------------------------------
st.write("")
st.markdown('<div class="section-title">Retribución al Fondo de Empleados</div>', unsafe_allow_html=True)

retribuciones_filtradas = []
for periodo in periodos_seleccionados:
    r = datos_por_periodo[periodo]["retribucion"]
    if not r.empty:
        r = r.copy()
        r["Periodo"] = periodo
        retribuciones_filtradas.append(r)

if not retribuciones_filtradas:
    st.caption("No hay productos de origen 'Fondo de Empleados' con ventas en los periodos seleccionados.")
else:
    retribucion_total = pd.concat(retribuciones_filtradas, ignore_index=True)
    total_a_retribuir = retribucion_total["Retribucion"].sum(skipna=True)

    kpi_col, tabla_col = st.columns([1, 2])
    render_kpi(kpi_col, "Total a Retribuir al Fondo", formato_pesos(total_a_retribuir), ACENTO_RETRIBUCION)

    with tabla_col:
        tabla_retribucion = retribucion_total[["Periodo", "Producto", "Categoria", "Total_Vendido", "Tasa", "Retribucion"]].rename(columns={
            "Categoria": "Categoría", "Total_Vendido": "Unidades", "Tasa": "Tasa por Unidad", "Retribucion": "Retribución",
        })
        tabla_retribucion["Categoría"] = tabla_retribucion["Categoría"].fillna("Sin categorizar")
        st.dataframe(
            tabla_retribucion.style.format({"Tasa por Unidad": lambda v: formato_pesos(v) if pd.notna(v) else "—",
                                             "Retribución": lambda v: formato_pesos(v) if pd.notna(v) else "—"}),
            use_container_width=True, hide_index=True,
        )
    if tabla_retribucion["Categoría"].eq("Sin categorizar").any():
        st.caption("Los productos 'Sin categorizar' no calzan con Camiseta/Chaqueta/Chompa/Gorra y no se incluyen en el total.")

# --------------------------------------------------------------
# TABLA COMPARATIVA POR PERIODO
# --------------------------------------------------------------
if hay_comparacion:
    st.write("")
    st.markdown('<div class="section-title">Resumen Comparativo por Periodo</div>', unsafe_allow_html=True)
    unidades_por_periodo = productos_filtrados.groupby("Periodo", as_index=False).size().rename(columns={"size": "Unidades Vendidas"})
    pedidos_por_periodo = df_filtrado.groupby("Periodo", as_index=False).size().rename(columns={"size": "Pedidos"})
    tabla = resumen_periodo.merge(unidades_por_periodo, on="Periodo", how="left").merge(pedidos_por_periodo, on="Periodo", how="left")
    tabla["Unidades Vendidas"] = tabla["Unidades Vendidas"].fillna(0).astype(int)
    tabla = tabla.rename(columns={"Consignacion_Neto": "Ingresos (neto)", "Valor_Mercancia": "Valor Mercancía"})
    tabla = tabla[["Periodo", "Ingresos (neto)", "Valor Mercancía", "Unidades Vendidas", "Pedidos"]]
    st.dataframe(
        tabla.style.format({"Ingresos (neto)": formato_pesos, "Valor Mercancía": formato_pesos}),
        use_container_width=True, hide_index=True,
    )

st.write("")

# --------------------------------------------------------------
# TABLA DE DATOS FILTRADOS
# --------------------------------------------------------------
st.markdown('<div class="section-title">Pedidos Filtrados</div>', unsafe_allow_html=True)
df_mostrar = df_filtrado[[
    "Periodo", "Fecha_Pago", "Medio_Pago", "Retribucion", "Descripcion",
    "Valor_Mercancia", "Envio", "Consignacion_Neto",
]].sort_values("Fecha_Pago", ascending=False).reset_index(drop=True)
df_mostrar = df_mostrar.rename(columns={
    "Fecha_Pago": "Fecha de Pago", "Medio_Pago": "Canal", "Retribucion": "Retribución",
    "Descripcion": "Descripción", "Valor_Mercancia": "Valor Mercancía", "Envio": "Envío",
    "Consignacion_Neto": "Consignación (neto)",
})
st.dataframe(df_mostrar, use_container_width=True, hide_index=True)

csv_descarga = df_mostrar.to_csv(index=False).encode("utf-8-sig")
st.download_button("Descargar pedidos filtrados (CSV)", data=csv_descarga, file_name="ventas_filtradas.csv", mime="text/csv")

# --------------------------------------------------------------
# PREGÚNTALE A LA IA
# --------------------------------------------------------------
st.write("")
st.markdown('<div class="section-title">Pregúntale a la IA sobre estos datos</div>', unsafe_allow_html=True)

if not GROQ_DISPONIBLE:
    st.caption("Instala el paquete 'groq' (ver requirements.txt) para activar esta sección.")
elif not groq_api_key:
    st.caption("Pega tu Groq API Key en la barra lateral para activar el chat.")
else:
    if "historial_chat_ia" not in st.session_state:
        st.session_state["historial_chat_ia"] = []

    for msg in st.session_state["historial_chat_ia"]:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    pregunta = st.chat_input("Ej: ¿cuál fue el producto más vendido? ¿por qué bajó la retribución este mes?")
    if pregunta:
        st.session_state["historial_chat_ia"].append({"role": "user", "content": pregunta})
        with st.chat_message("user"):
            st.markdown(pregunta)
        with st.chat_message("assistant"):
            with st.spinner("Pensando..."):
                try:
                    contexto = "\n\n".join(
                        construir_contexto_periodo(p, datos_por_periodo[p], incluir_pedidos=True)
                        for p in periodos_seleccionados
                    )
                    historial_para_ia = [
                        {"role": m["role"], "content": m["content"]}
                        for m in st.session_state["historial_chat_ia"][-6:-1]
                    ]
                    respuesta = preguntar_ia(pregunta, contexto, groq_api_key, historial=historial_para_ia)
                except Exception as e:
                    respuesta = f"No se pudo consultar la IA: {e}"
                st.markdown(respuesta)
        st.session_state["historial_chat_ia"].append({"role": "assistant", "content": respuesta})

    if st.session_state.get("historial_chat_ia"):
        if st.button("Limpiar conversación"):
            st.session_state["historial_chat_ia"] = []
            st.rerun()
